from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EXPORT_COLUMNS = [
    "LANÇAMENTO",
    "DESCRIÇÃO",
    "VENCIMENTO",
    "PAGAMENTO",
    "VALOR ORIGINAL",
    "VALOR",
    "PAGO BANCO",
    "STATUS",
]

BLUE = "4F81BD"
LIGHT_BLUE = "D9E6F2"
WHITE = "FFFFFF"
LIGHT_GREEN = "E8F8F1"
LIGHT_YELLOW = "FFF3CD"
LIGHT_RED = "F8D7DA"
BORDER = Side(style="thin", color="D9E2EF")


def parse_date_br(value) -> Optional[datetime]:
    if value is None or value == "" or str(value).lower() == "a revisar":
        return None
    if isinstance(value, datetime):
        return value
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


def parse_money(value) -> Optional[float]:
    if value is None or value == "" or str(value).lower() == "a revisar":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("R$", "").strip()
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def build_excel(df: pd.DataFrame, titulo: str = "CONTROLE DE DESPESAS") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle"

    # Cabeçalho institucional
    ws.merge_cells("A1:H1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, color="00457C", size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "IOT - INSTITUTO ORTOPEDICO MEDICAÇÃO"
    ws["A2"].font = Font(bold=True, color="00457C", size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:H3")
    ws["A3"] = "DESPESAS GERAIS"
    ws["A3"].font = Font(bold=True, color="00457C", size=11)
    ws["A3"].alignment = Alignment(horizontal="center")

    for row in range(1, 4):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    # Cabeçalho da tabela
    header_row = 5
    for idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=col_name)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=BORDER, bottom=BORDER, left=BORDER, right=BORDER)

    data_start = header_row + 1
    safe_df = df.copy()
    for col in EXPORT_COLUMNS:
        if col not in safe_df.columns:
            safe_df[col] = ""

    for r_idx, (_, row) in enumerate(safe_df[EXPORT_COLUMNS].iterrows(), start=data_start):
        for c_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=r_idx, column=c_idx)

            if col_name in ["LANÇAMENTO", "VENCIMENTO", "PAGAMENTO"]:
                parsed = parse_date_br(value)
                cell.value = parsed if parsed else value
                if parsed:
                    cell.number_format = "DD/MM/YYYY"
            elif col_name in ["VALOR ORIGINAL", "VALOR"]:
                parsed = parse_money(value)
                cell.value = parsed if parsed is not None else value
                if parsed is not None:
                    cell.number_format = 'R$ #,##0.00'
            else:
                cell.value = value

            cell.border = Border(top=BORDER, bottom=BORDER, left=BORDER, right=BORDER)
            cell.alignment = Alignment(vertical="center")

            status = str(row.get("STATUS", "")).lower()
            if "revis" in status:
                cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
            elif "erro" in status:
                cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
            elif "conferido" in status:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)

    total_row = data_start + len(safe_df)
    ws.cell(row=total_row, column=4, value="TOTAL")
    ws.cell(row=total_row, column=5, value=f"=SUM(E{data_start}:E{total_row-1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{data_start}:F{total_row-1})")
    ws.cell(row=total_row, column=7, value="JUROS")
    ws.cell(row=total_row, column=8, value=f"=F{total_row}-E{total_row}")

    for col in range(1, 9):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(bold=True, color="00457C")
        cell.border = Border(top=BORDER, bottom=BORDER, left=BORDER, right=BORDER)
        cell.alignment = Alignment(horizontal="center")
        if col in [5, 6, 8]:
            cell.number_format = 'R$ #,##0.00'

    widths = {
        "A": 16,
        "B": 42,
        "C": 16,
        "D": 16,
        "E": 18,
        "F": 18,
        "G": 24,
        "H": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:H{max(total_row - 1, header_row)}"
    ws.sheet_view.showGridLines = False

    # Aba de resumo com os indicadores principais.
    summary = wb.create_sheet("Resumo")
    summary["A1"] = "RESUMO DO CONTROLE DE DESPESAS"
    summary["A1"].font = Font(bold=True, color="00457C", size=14)
    metrics = [
        ("Total pago", f"=Controle!F{total_row}"),
        ("Valor original", f"=Controle!E{total_row}"),
        ("Juros", f"=Controle!H{total_row}"),
        ("Lançamentos", len(safe_df)),
        ("A revisar", f'=COUNTIF(Controle!H{data_start}:H{total_row-1},"*revisar*")+COUNTIF(Controle!H{data_start}:H{total_row-1},"*Erro*")+COUNTIF(Controle!H{data_start}:H{total_row-1},"*incompleto*")'),
    ]
    for idx, (label, value) in enumerate(metrics, start=3):
        summary.cell(row=idx, column=1, value=label)
        summary.cell(row=idx, column=2, value=value)
        summary.cell(row=idx, column=1).font = Font(bold=True)
        summary.cell(row=idx, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        summary.cell(row=idx, column=2).border = Border(top=BORDER, bottom=BORDER, left=BORDER, right=BORDER)
        if idx in [3, 4, 5]:
            summary.cell(row=idx, column=2).number_format = 'R$ #,##0.00'
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 20
    summary.sheet_view.showGridLines = False

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()
